import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Inspección SAG"

# Estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Metadata del documento
ws['A1'] = "PLANILLA DE DESCRIPCIÓN DE LOTE - INSPECCIÓN SAG"
ws['A1'].font = Font(bold=True, size=12)
ws.merge_cells('A1:P1')

ws['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}"
ws['A3'] = "N° Inspección: 2026-001"

# Filas vacías
ws.append([])

# Encabezados de columnas
headers = [
    "N° Folio", "CSG", "Productor", "Proceso", "Provincia Origen", "Comuna Origen",
    "CSP", "Provincia Pack.", "Comuna Pack.", "Especie", "Variedad Agronómica",
    "Variedad Comercial", "Fec.Pack", "SDP/Sector", "Cajas", "Total"
]

header_row = 5
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Datos de prueba - 3 folios con múltiples líneas
test_data = [
    # Folio G310150372 - 2 productores, 72 cajas
    {
        "folio": "G310150372",
        "total": 72,
        "lineas": [
            ("119185", "119185 SOC. E INMOBILIARIA PICHIDEGUA", "Packing", "Cachapoal", "PICHIDEGUA", "171972", "Cachapoal", "PICHIDEGUA", "LIMON", "GENOVA", "GENOVA", "01/06/2026", "", 22),
            ("119185", "", "", "", "", "", "", "", "", "", "", "", "", 15),
            ("119185", "", "", "", "", "", "", "", "", "", "", "", "", 20),
            ("140529", "140529 AGRICOLA LA CALERA", "Packing", "Cachapoal", "LA CALERA", "171972", "Cachapoal", "PICHIDEGUA", "LIMON", "GENOVA", "GENOVA", "01/06/2026", "", 15),
        ]
    },
    # Folio G310150373 - 1 productor, 48 cajas
    {
        "folio": "G310150373",
        "total": 48,
        "lineas": [
            ("142850", "142850 AGRICOLA SANTA ROSA", "Packing", "Colchagua", "SANTA CRUZ", "172100", "Colchagua", "SANTA CRUZ", "LIMON", "EUREKA", "EUREKA", "02/06/2026", "", 30),
            ("142850", "", "", "", "", "", "", "", "", "", "", "", "", 18),
        ]
    },
    # Folio G310150374 - 2 productores, 96 cajas
    {
        "folio": "G310150374",
        "total": 96,
        "lineas": [
            ("145000", "145000 FUNDO LOS MOLLES", "Packing", "Maule", "TALCA", "172200", "Maule", "TALCA", "LIMÓN", "FINO", "FINO", "03/06/2026", "A1", 40),
            ("145000", "", "", "", "", "", "", "", "", "", "", "", "", 20),
            ("150250", "150250 VIÑAS AGRICOLAS", "Packing", "Maule", "CURICO", "172200", "Maule", "TALCA", "LIMÓN", "FINO", "FINO", "03/06/2026", "B2", 36),
        ]
    }
]

# Escribir datos
current_row = header_row + 1
for folio_data in test_data:
    for idx, linea in enumerate(folio_data["lineas"]):
        row = [
            folio_data["folio"] if idx == 0 else "",  # N° Folio (solo en primera línea)
            linea[0],  # CSG
            linea[1],  # Productor
            linea[2],  # Proceso
            linea[3],  # Provincia Origen
            linea[4],  # Comuna Origen
            linea[5],  # CSP
            linea[6],  # Provincia Pack
            linea[7],  # Comuna Pack
            linea[8],  # Especie
            linea[9],  # Variedad Agronómica
            linea[10], # Variedad Comercial
            linea[11], # Fec.Pack
            linea[12], # SDP/Sector
            linea[13], # Cajas
            folio_data["total"] if idx == len(folio_data["lineas"]) - 1 else "",  # Total (solo en última línea)
        ]
        
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = value
            cell.border = border
            if col == 1:  # N° Folio
                cell.font = Font(bold=True)
            if col == 16:  # Total
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        current_row += 1

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 18
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 18
ws.column_dimensions['L'].width = 18
ws.column_dimensions['M'].width = 12
ws.column_dimensions['N'].width = 12
ws.column_dimensions['O'].width = 10
ws.column_dimensions['P'].width = 10

# Guardar
wb.save('c:/proyecto_pda/app/public/Planilla_Prueba_SAG.xlsx')
print("✓ Archivo de prueba creado: Planilla_Prueba_SAG.xlsx")
